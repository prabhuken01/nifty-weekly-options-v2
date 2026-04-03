#!/usr/bin/env python3
"""
Add Phase 2 roadmap to Excel: Historical data caching strategy
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = load_workbook('ShortStrangle_Dashboard.xlsx')

# Create Phase 2 Roadmap sheet
if 'Phase2_Roadmap' in wb.sheetnames:
    del wb['Phase2_Roadmap']
phase2_sheet = wb.create_sheet('Phase2_Roadmap', 1)

# Header
phase2_sheet['A1'] = 'Phase 2 Roadmap: Historical Data Caching Strategy'
phase2_sheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
phase2_sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
phase2_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
phase2_sheet.merge_cells('A1:E1')
phase2_sheet.row_dimensions[1].height = 25

# Problem statement
phase2_sheet['A3'] = 'Problem'
phase2_sheet['A3'].font = Font(bold=True, size=11, color='FFFFFF')
phase2_sheet['A3'].fill = PatternFill(start_color='E24B4A', end_color='E24B4A', fill_type='solid')
phase2_sheet['A4'] = 'Current implementation fetches option premiums from Kite API on every page load (Tab 2 live signal). Tab 1 backtest uses formula-based premium estimation instead of historical actual premiums. This approach has limitations:\n• Repeated API calls = rate limiting + latency\n• No historical premium archive = backtests inaccurate\n• Cannot validate strategy offline'
phase2_sheet['A4'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
phase2_sheet['A4'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
phase2_sheet.row_dimensions[4].height = 60

# Solution
phase2_sheet['A6'] = 'Solution'
phase2_sheet['A6'].font = Font(bold=True, size=11, color='FFFFFF')
phase2_sheet['A6'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
phase2_sheet['A7'] = 'Build a historical premium cache (Google Sheets or SQLite database) that stores:\n• Date, Time, Instrument, Strike, Expiry, Premium (CE), Premium (PE), IV, IVP\n• Updated weekly after expiry via separate scheduled job\n• Queried by backtest module to validate P&L assumptions'
phase2_sheet['A7'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
phase2_sheet['A7'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
phase2_sheet.row_dimensions[7].height = 60

# Implementation steps
phase2_sheet['A9'] = 'Implementation Steps'
phase2_sheet['A9'].font = Font(bold=True, size=11, color='FFFFFF')
phase2_sheet['A9'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

steps_data = [
    ['Step 1: Google Sheets Setup', 'Create "NIFTY_Premium_History" and "SENSEX_Premium_History" sheets in shared Google Drive. Each row: Date | Time | Strike | Expiry | IV | IVP | CE_Premium | PE_Premium | Notes. Share via gspread (Python library).', 'Week 1'],
    ['Step 2: Historical Data Backfill', 'Fetch last 12 weeks of NSE Bhavcopy option data via yfinance or nsepy. Parse CSV, extract premiums for ±2.5%/±3%/±3.5%/±4%/±4.5% strikes. Populate Google Sheets (1,000-1,500 rows per instrument).', 'Week 2'],
    ['Step 3: Scheduled Update Job', 'Create daily cron job (Cloud Functions or local scheduler) that:\n  a) Runs after NSE market close\n  b) Fetches today\'s option premiums from Kite API\n  c) Appends to Google Sheets\n  d) Updates IVP calculation\nKeep 52-week rolling window (delete rows older than 1 year).', 'Week 2-3'],
    ['Step 4: Backtest Integration', 'Update generate_backtest_pnl():\n  a) Query Google Sheets for date range matching lookback_m\n  b) For each offset, fetch mean/median premium from cache\n  c) Calculate win_rate and loss_per_trade from actual trade history\n  d) Compare formula-based P&L vs. actual historical P&L', 'Week 3'],
    ['Step 5: Validation & Dashboard', 'Add new Tab 4: "Historical Validation" showing:\n  - Backtest P&L (formula-based) vs. Actual Historical P&L (from cache)\n  - Win rate distribution by IV regime (LOW/MID/HIGH)\n  - Best performing offset + IV combination\n  - Confidence level: "High (12+ weeks data)" or "Low (< 4 weeks)"', 'Week 4'],
]

for idx, (step, description, timeline) in enumerate(steps_data, 10):
    phase2_sheet[f'A{idx}'] = step
    phase2_sheet[f'A{idx}'].font = Font(bold=True, size=10)
    phase2_sheet[f'A{idx}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    phase2_sheet[f'B{idx}'] = description
    phase2_sheet[f'B{idx}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    phase2_sheet[f'B{idx}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    phase2_sheet[f'C{idx}'] = timeline
    phase2_sheet[f'C{idx}'].font = Font(bold=True, italic=True)
    phase2_sheet[f'C{idx}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    phase2_sheet.row_dimensions[idx].height = 50

# Dependencies
phase2_sheet['A15'] = 'Dependencies & Tools'
phase2_sheet['A15'].font = Font(bold=True, size=11, color='FFFFFF')
phase2_sheet['A15'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

deps_data = [
    ['Google Sheets API + gspread (Python)', 'pip install gspread google-auth-oauthlib. Alternative: SQLite for local DB.'],
    ['NSE Bhavcopy fetch', 'yfinance or nsepy library. Alternative: Manual CSV download from NSE website.'],
    ['Cloud scheduler (optional)', 'Google Cloud Functions (free tier: 2M invocations/month). Alternative: Local cron job on always-on machine.'],
    ['Kite Connect API', 'Already integrated. Continue using for live quotes + daily scheduled updates.'],
]

for idx, (tool, note) in enumerate(deps_data, 17):
    phase2_sheet[f'A{idx}'] = tool
    phase2_sheet[f'A{idx}'].font = Font(bold=True)
    phase2_sheet[f'A{idx}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    phase2_sheet[f'B{idx}'] = note
    phase2_sheet[f'B{idx}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    phase2_sheet[f'B{idx}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    phase2_sheet.row_dimensions[idx].height = 30

# Set column widths
phase2_sheet.column_dimensions['A'].width = 25
phase2_sheet.column_dimensions['B'].width = 60
phase2_sheet.column_dimensions['C'].width = 12

# Add to Implementation_Status as pending
impl_sheet = wb['Implementation_Status']
next_row = impl_sheet.max_row + 2

impl_sheet[f'A{next_row}'] = 'Phase 2'
impl_sheet[f'B{next_row}'] = 'Historical Premium Cache'
impl_sheet[f'C{next_row}'] = 'PENDING'
impl_sheet[f'D{next_row}'] = 'Add Google Sheets/SQLite caching for option premiums; Build scheduled update job'
impl_sheet[f'E{next_row}'] = 'High'
impl_sheet[f'F{next_row}'] = 'Not started'
impl_sheet[f'G{next_row}'] = 'Enables accurate backtesting without repeated API calls; Required for Tab 4 validation'

# Style the new row
for col in range(1, 8):
    cell = impl_sheet.cell(row=next_row, column=col)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    if col == 3:  # Status column
        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        cell.font = Font(color='9C0006', bold=True)

impl_sheet.row_dimensions[next_row].height = 30

wb.save('ShortStrangle_Dashboard.xlsx')
print("[OK] Phase 2 Roadmap added to Excel")
print(f"[OK] New sheet: Phase2_Roadmap with 5-step implementation plan")
print(f"[OK] Implementation_Status updated with Phase 2 pending item")
