import pandas as pd
import openpyxl
from datetime import datetime

print("="*80)
print("BATCH 3: LUT RETRAINING IMPACT ANALYSIS")
print("="*80)

# Load IV impact data
print("\n[1/4] Loading IV impact analysis...")
iv_df = pd.read_csv('iv_impact_analysis.csv')
iv_df['old_band'] = None
iv_df['new_band'] = None

IV_BANDS = {
    '<13%': (0, 0.13),
    '13-15%': (0.13, 0.15),
    '15-18%': (0.15, 0.18),
    '18-22%': (0.18, 0.22),
    '>22%': (0.22, 999)
}

def get_band(iv):
    if pd.isna(iv):
        return None
    for band_name, (lo, hi) in IV_BANDS.items():
        if lo <= iv < hi:
            return band_name
    return None

iv_df['old_band'] = iv_df['old_iv'].apply(get_band)
iv_df['new_band'] = iv_df['new_iv'].apply(get_band)

print(f"  Loaded {len(iv_df)} dates")

# Load LUT from Excel
print("\n[2/4] Loading LUT from Nifty_Strategy_Selector_v2.xlsx...")
wb = openpyxl.load_workbook('Nifty_Strategy_Selector_v2.xlsx')
ws = wb.active

# Parse LUT
lut_rows = []
for row_idx in range(2, ws.max_row + 1):
    key_cell = ws.cell(row=row_idx, column=1).value
    strategy_cell = ws.cell(row=row_idx, column=2).value

    if not key_cell or not strategy_cell:
        continue

    lut_rows.append({
        'key': str(key_cell).strip(),
        'strategy': str(strategy_cell).strip()
    })

print(f"  Loaded {len(lut_rows)} LUT entries")

# Parse LUT keys to extract DTE, IV_BAND, TREND
lut_dict = {}
for entry in lut_rows:
    parts = entry['key'].split('|')
    if len(parts) >= 3:
        dte = parts[0].strip()
        iv_band = parts[1].strip()
        trend = parts[2].strip()
        key = f"{dte}|{iv_band}|{trend}"
        lut_dict[key] = entry['strategy']

print(f"  Parsed {len(lut_dict)} unique LUT entries")

# DTE mapping: Actual DTE to LUT row
def get_lut_dte_key(actual_dte):
    if actual_dte >= 4:
        return "T-4"
    elif actual_dte == 3:
        return "T-3"
    elif actual_dte == 2:
        return "T-2"
    elif actual_dte == 1:
        return "T-1"
    else:
        return "T-1"

# Simulate LUT lookups for old and new methods
print("\n[3/4] Simulating LUT lookups (old vs new)...")

# For simplicity, we'll assume TREND is NEUTRAL (most common)
TREND = "NEUTRAL"

impact_data = []

for idx, row in iv_df.iterrows():
    old_dte = row['old_dte']
    new_dte = row['new_dte']
    old_band = row['old_band']
    new_band = row['new_band']
    date = row['date']

    if old_dte is None or new_dte is None:
        continue
    if old_band is None or new_band is None:
        continue

    # Build LUT keys
    old_dte_key = get_lut_dte_key(old_dte)
    new_dte_key = get_lut_dte_key(new_dte)

    old_lut_key = f"{old_dte_key}|{old_band}|{TREND}"
    new_lut_key = f"{new_dte_key}|{new_band}|{TREND}"

    old_strategy = lut_dict.get(old_lut_key, 'NOT FOUND')
    new_strategy = lut_dict.get(new_lut_key, 'NOT FOUND')

    strategy_changed = old_strategy != new_strategy

    impact_data.append({
        'date': date,
        'old_dte': old_dte,
        'old_band': old_band,
        'old_dte_key': old_dte_key,
        'old_lut_key': old_lut_key,
        'old_strategy': old_strategy,
        'new_dte': new_dte,
        'new_band': new_band,
        'new_dte_key': new_dte_key,
        'new_lut_key': new_lut_key,
        'new_strategy': new_strategy,
        'strategy_changed': strategy_changed
    })

impact_df = pd.DataFrame(impact_data)

print(f"  Analyzed {len(impact_df)} dates")
print(f"  Strategy changes: {impact_df['strategy_changed'].sum()} dates ({100*impact_df['strategy_changed'].sum()/len(impact_df):.1f}%)")

# Summary analysis
print("\n[4/4] Generating summary...")

changed_df = impact_df[impact_df['strategy_changed'] == True]

print("\nLUT RETRAINING IMPACT ANALYSIS")
print("="*80)
print(f"\nTotal dates analyzed: {len(impact_df)}")
print(f"Dates with strategy change: {len(changed_df)} ({100*len(changed_df)/len(impact_df):.1f}%)")

if len(changed_df) > 0:
    print("\nStrategy Changes (breakdown by old->new):")
    change_counts = {}
    for idx, row in changed_df.iterrows():
        old_strat = row['old_strategy'][:20]
        new_strat = row['new_strategy'][:20]
        key = f"{old_strat} -> {new_strat}"
        change_counts[key] = change_counts.get(key, 0) + 1

    for key, count in sorted(change_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
        print(f"  {key}: {count} dates")

# Band shift summary
print("\nIV Band Distribution Changes:")
old_band_dist = impact_df['old_band'].value_counts().sort_index()
new_band_dist = impact_df['new_band'].value_counts().sort_index()

print("\nOld Method (nearest expiry):")
for band in ['<13%', '13-15%', '15-18%', '18-22%', '>22%']:
    count = old_band_dist.get(band, 0)
    pct = 100 * count / len(impact_df) if len(impact_df) > 0 else 0
    print(f"  {band:8s}: {count:3d} dates ({pct:5.1f}%)")

print("\nNew Method (DTE >= 2):")
for band in ['<13%', '13-15%', '15-18%', '18-22%', '>22%']:
    count = new_band_dist.get(band, 0)
    pct = 100 * count / len(impact_df) if len(impact_df) > 0 else 0
    print(f"  {band:8s}: {count:3d} dates ({pct:5.1f}%)")

# Recommendation
print("\n" + "="*80)
if len(changed_df) / len(impact_df) < 0.10:
    rec = "MINOR - < 10% strategy changes"
    action = "Can proceed with new IV; monitor alignment"
elif len(changed_df) / len(impact_df) < 0.25:
    rec = "MODERATE - 10-25% strategy changes"
    action = "Recommend LUT review; may need minor retraining"
else:
    rec = "MAJOR - > 25% strategy changes"
    action = "Recommend full LUT retraining with new IV method"

print(f"VERDICT: {rec}")
print(f"ACTION: {action}")
print("="*80)

# Save detailed report to CSV
impact_df.to_csv('LUT_Impact_Analysis.csv', index=False)
print(f"\nDetailed impact saved to: LUT_Impact_Analysis.csv")

# Create summary Excel
print("\nCreating summary Excel...")
wb_summary = openpyxl.Workbook()
ws_summary = wb_summary.active
ws_summary.title = "LUT Impact"

# Write summary
row = 1
ws_summary.cell(row=row, column=1).value = "LUT RETRAINING IMPACT ANALYSIS"
ws_summary.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True, size=12)
row += 2

ws_summary.cell(row=row, column=1).value = "Metric"
ws_summary.cell(row=row, column=2).value = "Value"
row += 1

ws_summary.cell(row=row, column=1).value = "Total dates analyzed"
ws_summary.cell(row=row, column=2).value = len(impact_df)
row += 1

ws_summary.cell(row=row, column=1).value = "Dates with strategy change"
ws_summary.cell(row=row, column=2).value = len(changed_df)
row += 1

ws_summary.cell(row=row, column=1).value = "% Strategy changes"
ws_summary.cell(row=row, column=2).value = f"{100*len(changed_df)/len(impact_df):.1f}%"
row += 2

ws_summary.cell(row=row, column=1).value = "VERDICT"
ws_summary.cell(row=row, column=2).value = rec
row += 1

ws_summary.cell(row=row, column=1).value = "RECOMMENDED ACTION"
ws_summary.cell(row=row, column=2).value = action
row += 2

ws_summary.cell(row=row, column=1).value = "Band Distribution (OLD)"
row += 1
for band in ['<13%', '13-15%', '15-18%', '18-22%', '>22%']:
    count = old_band_dist.get(band, 0)
    pct = 100 * count / len(impact_df) if len(impact_df) > 0 else 0
    ws_summary.cell(row=row, column=1).value = f"  {band}"
    ws_summary.cell(row=row, column=2).value = count
    ws_summary.cell(row=row, column=3).value = f"{pct:.1f}%"
    row += 1

row += 1
ws_summary.cell(row=row, column=1).value = "Band Distribution (NEW)"
row += 1
for band in ['<13%', '13-15%', '15-18%', '18-22%', '>22%']:
    count = new_band_dist.get(band, 0)
    pct = 100 * count / len(impact_df) if len(impact_df) > 0 else 0
    ws_summary.cell(row=row, column=1).value = f"  {band}"
    ws_summary.cell(row=row, column=2).value = count
    ws_summary.cell(row=row, column=3).value = f"{pct:.1f}%"
    row += 1

ws_summary.column_dimensions['A'].width = 40
ws_summary.column_dimensions['B'].width = 20
ws_summary.column_dimensions['C'].width = 20

wb_summary.save('LUT_Impact_Summary.xlsx')

print("\n" + "="*80)
print("BATCH 3 COMPLETE")
print("="*80)
print("\nOutputs:")
print("  - LUT_Impact_Analysis.csv (detailed impact for each date)")
print("  - LUT_Impact_Summary.xlsx (executive summary)")
print(f"\nNext: BATCH 4 will create IV Analysis tab for app")
