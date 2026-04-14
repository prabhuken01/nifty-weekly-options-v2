import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("="*80)
print("BATCH 2: IV IMPACT ANALYSIS EXCEL (Band Shift Detection)")
print("="*80)

# Load the CSV
print("\n[1/5] Loading iv_impact_analysis.csv...")
df = pd.read_csv('iv_impact_analysis.csv')
df['date'] = pd.to_datetime(df['date']).dt.date
print(f"  Loaded {len(df)} rows")

# IV bands (from app.py)
IV_BANDS = {
    '<13%': (0, 0.13),
    '13-15%': (0.13, 0.15),
    '15-18%': (0.15, 0.18),
    '18-22%': (0.18, 0.22),
    '>22%': (0.22, 999)
}

def get_band(iv):
    if iv is None or np.isnan(iv):
        return 'N/A'
    for band_name, (lo, hi) in IV_BANDS.items():
        if lo <= iv < hi:
            return band_name
    return 'N/A'

# Add band columns
print("\n[2/5] Computing IV bands (old vs new)...")
df['old_band'] = df['old_iv'].apply(get_band)
df['new_band'] = df['new_iv'].apply(get_band)
df['band_shift'] = df['old_band'] != df['new_band']

print(f"  Band shifts: {df['band_shift'].sum()}")

# Create Excel workbook
print("\n[3/5] Creating Excel workbook...")
wb = Workbook()
ws = wb.active
ws.title = "IV Impact"

# Define styles
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(size=10)
data_align_c = Alignment(horizontal='center', vertical='center')
data_align_r = Alignment(horizontal='right', vertical='center')

shift_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
shift_font = Font(size=10, bold=True)

border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Headers
headers = ['Date', 'Spot', 'ATM Strike',
           'Old Expiry', 'Old DTE', 'Old IV', 'Old Band',
           'New Expiry', 'New DTE', 'New IV', 'New Band',
           'IV Change', 'Band Shift?']

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# Data rows
print("\n[4/5] Populating data rows...")
for row_idx, row in df.iterrows():
    excel_row = row_idx + 2

    # Date
    cell = ws.cell(row=excel_row, column=1)
    cell.value = row['date']
    cell.alignment = data_align_c
    cell.border = border
    cell.font = data_font

    # Spot
    cell = ws.cell(row=excel_row, column=2)
    cell.value = round(float(row['spot']), 2) if pd.notna(row['spot']) else None
    cell.alignment = data_align_r
    cell.border = border
    cell.font = data_font

    # ATM Strike
    cell = ws.cell(row=excel_row, column=3)
    cell.value = int(row['atm_strike']) if pd.notna(row['atm_strike']) else None
    cell.alignment = data_align_r
    cell.border = border
    cell.font = data_font

    # Old Expiry
    cell = ws.cell(row=excel_row, column=4)
    cell.value = row['old_expiry']
    cell.alignment = data_align_c
    cell.border = border
    cell.font = data_font

    # Old DTE
    cell = ws.cell(row=excel_row, column=5)
    cell.value = int(row['old_dte']) if pd.notna(row['old_dte']) else None
    cell.alignment = data_align_r
    cell.border = border
    cell.font = data_font

    # Old IV
    cell = ws.cell(row=excel_row, column=6)
    cell.value = round(float(row['old_iv']), 4) if pd.notna(row['old_iv']) else None
    cell.alignment = data_align_r
    cell.border = border
    cell.font = data_font
    cell.number_format = '0.0000'

    # Old Band
    cell = ws.cell(row=excel_row, column=7)
    cell.value = row['old_band']
    cell.alignment = data_align_c
    cell.border = border
    cell.font = data_font

    # New Expiry
    cell = ws.cell(row=excel_row, column=8)
    cell.value = row['new_expiry']
    cell.alignment = data_align_c
    cell.border = border
    cell.font = data_font

    # New DTE
    cell = ws.cell(row=excel_row, column=9)
    cell.value = int(row['new_dte']) if pd.notna(row['new_dte']) else None
    cell.alignment = data_align_r
    cell.border = border
    cell.font = data_font

    # New IV
    cell = ws.cell(row=excel_row, column=10)
    cell.value = round(float(row['new_iv']), 4) if pd.notna(row['new_iv']) else None
    cell.alignment = data_align_r
    cell.border = border
    cell.font = data_font
    cell.number_format = '0.0000'

    # New Band
    cell = ws.cell(row=excel_row, column=11)
    cell.value = row['new_band']
    cell.alignment = data_align_c
    cell.border = border
    cell.font = data_font

    # IV Change
    cell = ws.cell(row=excel_row, column=12)
    if pd.notna(row['old_iv']) and pd.notna(row['new_iv']):
        cell.value = round(float(row['new_iv']) - float(row['old_iv']), 4)
        cell.number_format = '0.0000'
    else:
        cell.value = None
    cell.alignment = data_align_r
    cell.border = border
    cell.font = data_font

    # Band Shift?
    cell = ws.cell(row=excel_row, column=13)
    if row['band_shift']:
        cell.value = 'YES'
        cell.fill = shift_fill
        cell.font = shift_font
    else:
        cell.value = 'NO'
        cell.font = data_font
    cell.alignment = data_align_c
    cell.border = border

# Adjust column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 10
ws.column_dimensions['L'].width = 12
ws.column_dimensions['M'].width = 12

# Freeze first row
ws.freeze_panes = 'A2'

# Create Summary sheet
print("\n[5/5] Creating Summary sheet...")
ws_summary = wb.create_sheet('Summary')

summary_data = []
summary_data.append(['BATCH 2: IV IMPACT ANALYSIS SUMMARY', '', '', ''])
summary_data.append(['', '', '', ''])
summary_data.append(['Metric', 'Value', '', ''])
summary_data.append(['Total dates analyzed', len(df), '', ''])
summary_data.append(['Dates with band shift', df['band_shift'].sum(), '', f"{100*df['band_shift'].sum()/len(df):.1f}% of dates"])

# IV change stats
both_valid = df[(df['old_iv'].notna()) & (df['new_iv'].notna())].copy()
if len(both_valid) > 0:
    both_valid['iv_change'] = both_valid['new_iv'] - both_valid['old_iv']
    summary_data.append(['', '', '', ''])
    summary_data.append(['IV Change Statistics (both old & new available)', '', '', ''])
    summary_data.append(['Dates with data', len(both_valid), '', ''])
    summary_data.append(['Mean IV change', f"{both_valid['iv_change'].mean():.4f}", '', ''])
    summary_data.append(['Median IV change', f"{both_valid['iv_change'].median():.4f}", '', ''])
    summary_data.append(['Std Dev', f"{both_valid['iv_change'].std():.4f}", '', ''])
    summary_data.append(['Min change', f"{both_valid['iv_change'].min():.4f}", '', ''])
    summary_data.append(['Max change', f"{both_valid['iv_change'].max():.4f}", '', ''])

# Band distribution
summary_data.append(['', '', '', ''])
summary_data.append(['Band Distribution (OLD METHOD)', 'Count', '% of Total', ''])
for band in ['<13%', '13-15%', '15-18%', '18-22%', '>22%', 'N/A']:
    count = (df['old_band'] == band).sum()
    pct = 100 * count / len(df) if len(df) > 0 else 0
    summary_data.append([f"  {band}", count, f"{pct:.1f}%", ''])

summary_data.append(['', '', '', ''])
summary_data.append(['Band Distribution (NEW METHOD - DTE>=2)', 'Count', '% of Total', ''])
for band in ['<13%', '13-15%', '15-18%', '18-22%', '>22%', 'N/A']:
    count = (df['new_band'] == band).sum()
    pct = 100 * count / len(df) if len(df) > 0 else 0
    summary_data.append([f"  {band}", count, f"{pct:.1f}%", ''])

# Write summary
for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        cell.font = Font(size=10)

        if row_idx in [1, 3, 20, 27]:
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        elif row_idx == 5:
            cell.font = Font(bold=True, size=10)

ws_summary.column_dimensions['A'].width = 50
ws_summary.column_dimensions['B'].width = 15
ws_summary.column_dimensions['C'].width = 15
ws_summary.column_dimensions['D'].width = 30

# Save workbook
print("\n  Saving workbook...")
wb.save('IV_Impact_Analysis.xlsx')

print("\n" + "="*80)
print("BATCH 2 COMPLETE")
print("="*80)
print("\nOutput: IV_Impact_Analysis.xlsx")
print("  - Sheet 1 'IV Impact': All 371 dates with band analysis")
print("  - Sheet 2 'Summary': Statistics & distribution tables")
print(f"\nBand shifts: {df['band_shift'].sum()} dates ({100*df['band_shift'].sum()/len(df):.1f}%)")
print(f"Mean IV change: {both_valid['iv_change'].mean():.4f}")
print("\nNext: BATCH 3 will analyze LUT impact from band shifts")
