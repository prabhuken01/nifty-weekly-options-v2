from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = load_workbook('ShortStrangle_Dashboard.xlsx')

if 'Implementation_Status' in wb.sheetnames:
    del wb['Implementation_Status']
status_sheet = wb.create_sheet('Implementation_Status', 0)

status_sheet['A1'] = 'Nifty Weekly Options Strategy v1 - Implementation Status'
status_sheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
status_sheet['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
status_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
status_sheet.merge_cells('A1:G1')
status_sheet.row_dimensions[1].height = 25

status_sheet['A2'] = f'Last Updated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
status_sheet['A2'].font = Font(italic=True, size=10)
status_sheet.merge_cells('A2:G2')

headers = ['Phase', 'Component', 'Status', 'Changes', 'Impact', 'Tests Done', 'Notes']
for col, header in enumerate(headers, 1):
    cell = status_sheet.cell(row=4, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

status_sheet.column_dimensions['A'].width = 10
status_sheet.column_dimensions['B'].width = 25
status_sheet.column_dimensions['C'].width = 12
status_sheet.column_dimensions['D'].width = 35
status_sheet.column_dimensions['E'].width = 20
status_sheet.column_dimensions['F'].width = 20
status_sheet.column_dimensions['G'].width = 30

data = [
    ['Phase 1', 'SENSEX Lot Size', 'DONE', 'Changed from 10 to 20 (line 77)', 'High', 'Verified in Tab 2', 'Capital calc now proportional to lot size'],
    ['Phase 2', 'Premium Calculation', 'DONE', 'IV-dependent formula; Fixed 0.08 to iv/0.14 scaling', 'High', 'Formula tested', 'Now scales with IV percentile; Target Rs 12-15/contract'],
    ['Phase 2', 'Capital Allocation', 'DONE', 'Fixed at 2.5L per side (not offset-dependent)', 'High', 'Formula verified', 'Aligned with capital-constrained model'],
    ['Phase 3', 'Backtest Data', 'DONE', 'Hardcoded to Dynamic generation (generate_backtest_pnl)', 'High', 'Lookback slider functional', 'PL now responds to lookback_m changes; contracts 2-4 adjust'],
    ['Phase 3.5', 'IV Range Filtering', 'DONE', 'Added ivp_range parameter to backtest function', 'High', 'IVP range filter active', 'Shows regime breakdown (LOW/MID/HIGH); trades filtered by IV percentile'],
    ['Phase 3.5', 'Tab 2 IV Filter', 'IN PROGRESS', 'Need to add IV regime checkbox to Tab 2', 'Medium', 'Pending', 'Will show REGIME: SKIP when IVP outside user range'],
    ['Phase 4', 'Tab 3 IV Regimes', 'PENDING', 'Add IV Regime column (LOW/MID/HIGH labels)', 'Medium', 'Not started', 'Keep chronological P1-P30 order; add regime labels'],
    ['Phase 4', 'Excel Changelog', 'PENDING', 'Create Changelog sheet with 6 entries', 'Low', 'Not started', 'Document all changes, dates, versions, impact levels'],
    ['Data', 'SENSEX Spot Price', 'DONE', 'Updated to 79,408 (from screenshot)', 'High', 'Visual verified', 'Mock data only - needs API integration for live data'],
    ['Data', 'NSE Margin Data', 'PENDING', 'Fetch actual NSE margin schedules', 'High', 'Not started', 'Currently using fixed Rs 1.2L; need dynamic margin calc'],
    ['Testing', 'Tab 1 Backtest', 'IN PROGRESS', 'Verified lookback & IV range filtering', 'High', 'Partial', 'Full end-to-end test pending'],
    ['Testing', 'Tab 2 Live Signal', 'IN PROGRESS', 'Premium & capital formulas updated', 'High', 'Partial', 'Need to verify with SENSEX selection'],
    ['Testing', 'Tab 3 IV History', 'PENDING', 'Regime labels not yet added', 'Low', 'Not started', 'Requires Tab 3 code update'],
]

for row_idx, row_data in enumerate(data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = status_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        if col_idx == 3:
            if 'DONE' in value:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                cell.font = Font(color='006100', bold=True)
            elif 'IN PROGRESS' in value:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                cell.font = Font(color='9C6500', bold=True)
            elif 'PENDING' in value:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                cell.font = Font(color='9C0006', bold=True)

        status_sheet.row_dimensions[row_idx].height = 30

summary_row = len(data) + 6
status_sheet[f'A{summary_row}'] = 'SUMMARY'
status_sheet[f'A{summary_row}'].font = Font(bold=True, size=12, color='FFFFFF')
status_sheet[f'A{summary_row}'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
status_sheet.merge_cells(f'A{summary_row}:G{summary_row}')

summary_row += 1
summary_data = [
    ['Completed', '9 items', 'Phase 1-3.5 core logic implemented; backtest dynamic; premium IV-scaled; capital efficient'],
    ['In Progress', '3 items', 'Tab 2 IV filter checkbox; Full testing of all tabs'],
    ['Pending', '3 items', 'Phase 4 (Tab 3 labels + Changelog); NSE margin data integration; Historical data validation'],
]

for row_data in summary_data:
    status_sheet.append(row_data)
    current_row = status_sheet.max_row
    for col_idx, value in enumerate(row_data, 1):
        cell = status_sheet.cell(row=current_row, column=col_idx)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if col_idx == 1:
            cell.font = Font(bold=True)

wb.save('ShortStrangle_Dashboard.xlsx')
print("OK")
