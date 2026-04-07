#!/usr/bin/env python3
"""Create Excel documentation for Nifty options data pipeline"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = Workbook()
wb.remove(wb.active)

# ============================================================================
# SHEET 1: INSTRUCTIONS
# ============================================================================
ws1 = wb.create_sheet("Instructions", 0)
ws1.column_dimensions['A'].width = 45
ws1.column_dimensions['B'].width = 70

header_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)

ws1['A1'] = "NIFTY WEEKLY OPTIONS - DATA PIPELINE SETUP"
ws1['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws1['A1'].fill = header_fill
ws1.merge_cells('A1:B1')

row = 3
instructions = [
    ("STEP 1: CREATE SCHEMA (Run ONCE)",
     "python setup_schema_enhanced.py",
     "Creates database with:\n- option_bars_daily (6-month history)\n- option_bars_minute (future weeklies)\n- data_metadata (tracking)\nTime: 2 minutes"),

    ("STEP 2: BACKFILL 6 MONTHS (Run ONCE)",
     "python kite_fetch_nifty_daily_6months.py",
     "Downloads 6 months of NIFTY daily options.\nFor backtesting historical performance.\nTime: 30 minutes"),

    ("STEP 3: DAILY MINUTE DATA (Run EVERY DAY @ 4 PM)",
     "python kite_fetch_nifty_minute_daily.py",
     "Fetches minute-level data for current/next week.\nFor live trading and real-time Greeks.\nTime: 2 minutes"),

    ("FALLBACK: Manual CSV",
     "python nifty_bhavcopy_manual.py",
     "If Kite API unavailable:\n1. Download CSV from NSE\n2. Save to ./bhavcopies/ folder\n3. Run this script"),
]

for step, command, desc in instructions:
    ws1[f'A{row}'] = step
    ws1[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws1[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws1.merge_cells(f'A{row}:B{row}')
    row += 1

    ws1[f'A{row}'] = "Command:"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = command
    ws1[f'B{row}'].font = Font(bold=True, size=10)
    row += 1

    ws1[f'A{row}'] = "Details:"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = desc
    ws1[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 2

# ============================================================================
# SHEET 2: SCHEMA
# ============================================================================
ws2 = wb.create_sheet("Database Schema", 1)
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 45

# Header
for col, header in enumerate(['Column Name', 'Data Type', 'Nullable', 'Description'], 1):
    cell = ws2.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Table 1
ws2['A2'] = "TABLE: option_bars_daily (6-Month History)"
ws2['A2'].font = Font(bold=True, size=11, color="FFFFFF")
ws2['A2'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws2.merge_cells('A2:D2')

cols_daily = [
    ('id', 'SERIAL', 'No', 'Primary key'),
    ('timestamp', 'TIMESTAMP', 'No', 'Trading time (15:30 IST)'),
    ('symbol', 'VARCHAR(10)', 'No', 'NIFTY or SENSEX'),
    ('strike', 'DECIMAL(10,2)', 'No', 'Strike price'),
    ('option_type', 'VARCHAR(2)', 'No', 'CE (Call) or PE (Put)'),
    ('expiry', 'DATE', 'No', 'Expiry date (YYYY-MM-DD)'),
    ('open', 'DECIMAL(10,4)', 'Yes', 'Opening price'),
    ('high', 'DECIMAL(10,4)', 'Yes', 'High price of day'),
    ('low', 'DECIMAL(10,4)', 'Yes', 'Low price of day'),
    ('close', 'DECIMAL(10,4)', 'No', 'Closing price (IMPORTANT)'),
    ('volume', 'BIGINT', 'Yes', 'Contracts traded'),
    ('open_interest', 'BIGINT', 'Yes', 'Open interest'),
    ('iv', 'DECIMAL(10,4)', 'Yes', 'Implied volatility (%)'),
    ('delta', 'DECIMAL(10,4)', 'Yes', 'Delta: Price sensitivity'),
    ('gamma', 'DECIMAL(10,4)', 'Yes', 'Gamma: Delta acceleration'),
    ('theta', 'DECIMAL(10,4)', 'Yes', 'Theta: Daily time decay'),
    ('vega', 'DECIMAL(10,4)', 'Yes', 'Vega: IV sensitivity'),
    ('rho', 'DECIMAL(10,4)', 'Yes', 'Rho: Rate sensitivity'),
]

row = 3
for col_name, col_type, nullable, desc in cols_daily:
    ws2[f'A{row}'] = col_name
    ws2[f'B{row}'] = col_type
    ws2[f'C{row}'] = nullable
    ws2[f'D{row}'] = desc
    row += 1

row += 2

# Table 2
ws2[f'A{row}'] = "TABLE: option_bars_minute (Minute Candles)"
ws2[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
ws2[f'A{row}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws2.merge_cells(f'A{row}:D{row}')
row += 1

ws2[f'A{row}'] = "Same columns as option_bars_daily, for 1-minute interval data"
ws2.merge_cells(f'A{row}:D{row}')
row += 2

# Table 3
ws2[f'A{row}'] = "TABLE: data_metadata (Update Tracking)"
ws2[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
ws2[f'A{row}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws2.merge_cells(f'A{row}:D{row}')
row += 1

cols_meta = [
    ('id', 'SERIAL', 'No', 'Primary key'),
    ('table_name', 'VARCHAR(50)', 'No', 'daily or minute'),
    ('last_update', 'TIMESTAMP', 'No', 'Last fetch time'),
    ('data_type', 'VARCHAR(20)', 'No', 'DAILY or MINUTE'),
    ('source', 'VARCHAR(100)', 'No', 'Kite, NSE, Dhan, Manual'),
    ('records_count', 'BIGINT', 'Yes', 'Total rows in table'),
    ('date_range_from', 'DATE', 'Yes', 'Data start date'),
    ('date_range_to', 'DATE', 'Yes', 'Data end date'),
]

for col_name, col_type, nullable, desc in cols_meta:
    ws2[f'A{row}'] = col_name
    ws2[f'B{row}'] = col_type
    ws2[f'C{row}'] = nullable
    ws2[f'D{row}'] = desc
    row += 1

# ============================================================================
# SHEET 3: DATA FORMATS
# ============================================================================
ws3 = wb.create_sheet("Data Formats", 2)
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 50

# Daily format
ws3['A1'] = "DAILY DATA FORMAT (Option close prices at 15:30 IST)"
ws3['A1'].font = Font(bold=True, size=12, color="FFFFFF")
ws3['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws3.merge_cells('A1:C1')

headers = ['Field', 'Example Value', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=2, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

daily_examples = [
    ('timestamp', '2026-04-05 15:30:00', 'Market close time'),
    ('symbol', 'NIFTY', 'Index name'),
    ('strike', '24500', 'Strike price'),
    ('option_type', 'CE', 'Call or Put'),
    ('expiry', '2026-04-13', 'YYYY-MM-DD format'),
    ('close', '1270.50', 'Close price in INR'),
    ('iv', '28.5', 'Volatility in % (0-100)'),
    ('delta', '0.65', 'Range: -1 to +1'),
    ('theta', '-0.025', 'Time decay per day'),
]

row = 3
for field, example, note in daily_examples:
    ws3[f'A{row}'] = field
    ws3[f'B{row}'] = example
    ws3[f'C{row}'] = note
    row += 1

row += 2

# Minute format
ws3[f'A{row}'] = "MINUTE DATA FORMAT (1-minute candles for current/next week)"
ws3[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
ws3[f'A{row}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws3.merge_cells(f'A{row}:C{row}')
row += 1

for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

row += 1
minute_examples = [
    ('timestamp', '2026-04-05 14:05:00', '1-minute bar timestamp'),
    ('symbol', 'NIFTY', 'Index name'),
    ('strike', '24500', 'Strike price'),
    ('option_type', 'PE', 'Call or Put'),
    ('expiry', '2026-04-13', 'Current or next week'),
    ('close', '147.25', 'Close price in INR'),
    ('volume', '500', 'Contracts traded'),
    ('iv', '28.7', 'Updated every minute'),
    ('delta', '0.63', 'Updated every minute'),
]

for field, example, note in minute_examples:
    ws3[f'A{row}'] = field
    ws3[f'B{row}'] = example
    ws3[f'C{row}'] = note
    row += 1

# ============================================================================
# SHEET 4: WORKFLOW
# ============================================================================
ws4 = wb.create_sheet("Workflow", 3)
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 30

ws4['A1'] = "EXECUTION WORKFLOW"
ws4['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws4['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws4.merge_cells('A1:C1')

headers = ['Step', 'File to Run', 'Purpose & Notes']
for col, header in enumerate(headers, 1):
    cell = ws4.cell(row=2, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

workflow = [
    ('[ONCE]\nStep 1', 'setup_schema_enhanced.py', 'Creates 3 tables:\n- option_bars_daily\n- option_bars_minute\n- data_metadata\nTime: 2 minutes'),

    ('[ONCE]\nStep 2', 'kite_fetch_nifty_daily_6months.py', 'Backfill 6 months of daily data.\nFor historical backtesting.\nRun immediately after Step 1.\nTime: 30 minutes'),

    ('[DAILY]\nStep 3', 'kite_fetch_nifty_minute_daily.py', 'Run every day at 4 PM IST.\nFetches minute data for current\nand next week weeklies.\nTime: 2 minutes'),

    ('[IF NEEDED]\nFallback', 'nifty_bhavcopy_manual.py', 'Use if Kite API unavailable.\nDownload CSV from NSE first.\nSave to ./bhavcopies/ folder.'),
]

row = 3
for step, file, purpose in workflow:
    ws4[f'A{row}'] = step
    ws4[f'B{row}'] = file
    ws4[f'B{row}'].font = Font(bold=True, size=10)
    ws4[f'C{row}'] = purpose
    ws4[f'C{row}'].alignment = Alignment(wrap_text=True)
    row += 2

row += 1
ws4[f'A{row}'] = "Key Points:"
ws4[f'A{row}'].font = Font(bold=True, size=11)
row += 1

notes = [
    "✓ Step 1 must complete before Steps 2 & 3",
    "✓ Step 2 backfills historical data (one-time, 30 mins)",
    "✓ Step 3 runs daily at 4 PM for current/next week data",
    "✓ Data automatically stored in correct tables",
    "✓ metadata table auto-updated after each run",
    "✓ Kite API authenticated (already logged in)",
    "✓ All fallback methods configured",
]

for note in notes:
    ws4[f'A{row}'] = note
    row += 1

# Save
output_file = "NIFTY_Options_Data_Pipeline.xlsx"
wb.save(output_file)
print(f"[OK] Excel file created: {output_file}")
print(f"[OK] File contains 4 sheets:")
print(f"     1. Instructions - Setup guide")
print(f"     2. Database Schema - Column details")
print(f"     3. Data Formats - Example data")
print(f"     4. Workflow - Execution sequence")
